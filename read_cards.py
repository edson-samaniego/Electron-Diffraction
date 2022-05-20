import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

def DRX(muestras):
     DF=[]
     for tarjeta in muestras:
          records = []
          header = None
          parse = False
          with open(tarjeta, encoding = 'ISO-8859-1') as source:
               for raw in source:
                    clean = raw.encode('ascii', 'ignore').decode()
                    clean = clean.replace(' [', '[')
                    line = re.sub(r'[^a-zA-Z0-9. ]', '', clean)
                    line = line.replace('par', '')
                    line = line.replace('c5', 'A')          
                    line = line.replace('ulf2', '')
                    line = line.replace('ulnone', '')
                    if not parse:
                         if 'Peak list' in line:
                              parse = True
                    else:
                         if 'Structure' in line:
                              parse = False
                         else:
                              if 'No.' in line:
                                   header = line.split()
                              elif header is not None:
                                   fields = line.split()
                                   if len(fields) == len(header):
                                        ints = [ int(f) for f in fields[:4] ]
                                        floats = [ float(f) for f in fields[4:] ]
                                        records.append(ints + floats)
          df = pd.DataFrame.from_records(records, columns = header)
          DF.append(df)    
     return(df)

def PTC_Lab(muestra):
     wb = load_workbook('tarjetas_cif.xlsx')
     sheet = wb[muestra]
     row_count = sheet.max_row
     column_count = sheet.max_column
     cnt=0
     for i in range(1, row_count + 1):
         fila=[]    
         for j in range(1, column_count + 1):
             data = sheet.cell(row=i, column=j).value
             fila.append(data)
         if cnt==0:
           df = pd.DataFrame(columns=fila)
         else:
             new= pd.Series(fila, index=df.columns)
             df = df.append(new,ignore_index=True) 
         cnt=cnt+1
     return(df)











     
